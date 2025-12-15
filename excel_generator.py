"""Генератор Excel-файлов с извещениями на оплату."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import PaymentNotice
from date_time_utils import DateTimeHandler


class ExcelGenerator:
    """Генератор Excel-файлов."""

    def __init__(self):
        self._date_handler = DateTimeHandler()

    def generate_payment_notice(self, notice: PaymentNotice, file_path: str) -> None:
        """Генерирует Excel-файл с извещением на оплату."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Извещение на оплату"

        # Стили
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        label_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Заголовок
        ws.merge_cells('A1:E1')
        ws['A1'].value = "ИЗВЕЩЕНИЕ НА ОПЛАТУ"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Информация о лицевом счете
        row = 3
        for label, value in [
            ("Лицевой счет:", notice.account.account_number),
            ("ФИО:", notice.account.full_name),
            ("Адрес:", notice.account.get_address(notice.street)),
            ("Период:", f"{self._date_handler.get_month_name(notice.period_month)} {notice.period_year}"),
            ("Дата формирования:", self._date_handler.format_date(self._date_handler.get_current_date()))
        ]:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = label_font
            row += 1

        # Таблица начислений
        row += 1
        headers = ["№", "Услуга", "Количество", "Тариф", "Сумма"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Данные начислений
        row += 1
        for idx, (charge, service) in enumerate(notice.charges, start=1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=service.name).border = border
            ws.cell(row=row, column=3, value=charge.quantity).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=4, value=service.tariff).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            cost = service.calculate_cost(charge.quantity)
            ws.cell(row=row, column=5, value=cost).border = border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            row += 1

        # Итого
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = "ИТОГО К ОПЛАТЕ:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{row}'].border = border
        ws[f'E{row}'].value = notice.total_amount
        ws[f'E{row}'].font = Font(bold=True, size=12)
        ws[f'E{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'E{row}'].border = border
        ws[f'E{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Ширина столбцов
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        wb.save(file_path)

